# -*- coding: utf-8 -*-
"""
Created on Mon Jan 25 13:17:05 2021

@author: Samuel Solomon
"""

# Import Modules
# Basic Modules
import re
import time
import os
import math
import sys
import numpy as np
# Read/Write to Excel
import openpyxl as xl
from openpyxl.styles.colors import Color
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
# Input Data from Arduino
import serial
import serial.tools.list_ports
# Peak Detection
import scipy
import scipy.signal
from scipy.signal import lfilter
# Plotting
import matplotlib.pyplot as plt
import matplotlib.animation as manimation
# Neural Network
import neuralNetwork as NeuralNet
from sklearn.model_selection import train_test_split


# --------------------------------------------------------------------------- #
# ------------------ User Can Edit (Global Variables) ----------------------- #


# Set Global Variable 'data': The Input Data with Voltages/Time of Run
data = dict(time_ms=[], Channel1=[], Channel2=[], Channel3=[], Channel4=[])  

# Specify Figure Asthetics
peakCurrentRightColorOrder = {
    0: "tab:red",
    1: "tab:purple",
    2: "tab:orange",
    3: "tab:pink",
    4: "tab:brown",
    5: "tab:green",
    6: "tab:gray",
    7: "tab:cyan",
    }

# use ggplot style for more sophisticated visuals
plt.style.use('ggplot')

# Specify OpenPyxl Asthetics
openpyxlColors = {
    0: 45,
    1: 46,
    2: 47,
    3: 49,
    4: 50,
    5: 51,
    6: 55,
    }


# --------------------------------------------------------------------------- #
# ------------ Plot Variables user Can Edit (Global Variables) -------------- #
# High Pass Filter Parameters
f1 = 100; f3 = 50;
Rp = 0.1; Rs = 30;
samplingFreq = 1000
# Root Mean Squared (RMS) Parameters
rmsWindow = 250; stepSize = 8;

# Data Collection Parameters
RMSDataList = {};
xTopGrouping = {};
featureSetGrouping = {};
mapPeakLocToYPos = {};
peakDetectionBufferSize = 500
featureDefinitionBuffer = 50
maxPeakSep = 100   # Seperation that Defines a New Group
badInitialHighPass = max(rmsWindow+stepSize, 5000)  # Must be > rmsWindow + stepSize; Current Experimental lowest: xWidth*0.4 (Changes with xWidth)

    
# Specify Figure aesthetics
plt.ion()
figWidth = 14; figHeight = 10; # 28,16
fig, ax = plt.subplots(4, 2, sharey=False, sharex = 'col', figsize=(figWidth, figHeight))

#Initialize Movie Writer for Plots
movieSpeed = 20 # Speed of the Saved Movie (Frames per Second)
FFMpegWriter = manimation.writers['ffmpeg']
metadata = dict(title="EMG Channel Signals", artist='Matplotlib', comment='Movie support!')
writer = FFMpegWriter(fps=movieSpeed, metadata=metadata)

# Plot the Raw Data
xWidth = 2000
moveDataFinger = 200
xWidthPeaks = 10000
xLimLow = 0; xLimHigh = xLimLow + xWidth;
yLimLow = 0; yLimHigh = 5; 
movieGraphChannelListRaw = []
channelListRaw = []
numChannels = 4
for channelNum in range(numChannels):
    # Create Plots
    channelListRaw.append(ax[channelNum, 0])
    
    # Generate Plot
    movieGraphChannelListRaw.append(channelListRaw[channelNum].plot([], [], '-', c="tab:red", linewidth=1, alpha = 0.65)[0])
    
    # Set Figure Limits
    channelListRaw[channelNum].set_xlim(xLimLow, xLimHigh)
    channelListRaw[channelNum].set_ylim(yLimLow, yLimHigh)
    # Label Axis + Add Title
    channelListRaw[channelNum].set_title("EMG Signal in Channel " + str(channelNum + 1))
    channelListRaw[channelNum].set_xlabel("EMG Data Points")
    channelListRaw[channelNum].set_ylabel("EMG Signal (Volts)")
    
# Create the Peak Data Plot
yLimitHighFiltered = 0.5;
channelListFiltered = [] 
# Plot the Peak Data
movieGraphChannelListFiltered = []
movieGraphChannelTopPeaksList = []
movieGraphChannelBottomPeaksList = []
for channelNum in range(numChannels):
    # Create Plots
    channelListFiltered.append(ax[channelNum, 1])
    
    # Plot RMS Peaks
    movieGraphChannelListFiltered.append(channelListFiltered[channelNum].plot([], [], '-', c="tab:red", linewidth=1, alpha = 0.65)[0])
    # Plot Top Peaks
    movieGraphChannelTopPeaksList.append({})
    
    # Set Figure Limits
    channelListFiltered[channelNum].set_ylim(yLimLow, yLimitHighFiltered)
    # Label Axis + Add Title
    channelListFiltered[channelNum].set_title("Filtered EMG Signal in Channel " + str(channelNum + 1))
    channelListFiltered[channelNum].set_xlabel("Root Mean Squared Data Point")
    channelListFiltered[channelNum].set_ylabel("Filtered Signal (Volts)")
    
    # Hold Analysis Values
    RMSDataList[channelNum] = []
    mapPeakLocToYPos[channelNum] = {}
    xTopGrouping[channelNum] = {1:[]}
    featureSetGrouping[channelNum] = {1:[]}
    

# Tighten Figure White Space (Must be After wW Add Fig Info)
fig.tight_layout(pad=2.0)
plt.show()

# Create Output File Directory to Save Data
outputData = "../Output Data/"
os.makedirs(outputData, exist_ok=True)

# Keep Track of Robotic Movements
lastPeakAnalyzed = 0
currentGroupNum = -1

# Set up connection
HANDSHAKE = 0
VOLTAGE_REQUEST = 1
ON_REQUEST = 2;
STREAM = 3;
READ_DAQ_DELAY = 4;


# --------------------------------------------------------------------------- #
# ----------------- Stream Data from Arduino Can Edit ----------------------- #


def find_arduino(port=None):
    """Get the name of the port that is connected to Arduino."""
    if port is None:
        ports = serial.tools.list_ports.comports()
        for p in ports:
            if p.manufacturer is not None and ("Arduino" in p.manufacturer or "Microsoft" in p.manufacturer):
                port = p.device
    return port


def handshake_arduino(
    arduino, sleep_time=1, print_handshake_message=False, handshake_code=0
):
    """Make sure connection is established by sending
    and receiving bytes."""
    # Close and reopen
    arduino.close()
    arduino.open()

    # Chill out while everything gets set
    time.sleep(sleep_time)

    # Set a long timeout to complete handshake
    timeout = arduino.timeout
    arduino.timeout = 2

    # Read and discard everything that may be in the input buffer
    _ = arduino.read_all()

    # Send request to Arduino
    arduino.write(bytes([handshake_code]))

    # Read in what Arduino sent
    handshake_message = arduino.read_until()

    # Send and receive request again
    arduino.write(bytes([handshake_code]))
    handshake_message = arduino.read_until()

    # Print the handshake message, if desired
    if print_handshake_message:
        print("Handshake message: " + handshake_message.decode())

    # Reset the timeout
    arduino.timeout = timeout


def read_all(ser, read_buffer=b"", **args):
    """Read all available bytes from the serial port
    and append to the read buffer.

    Parameters
    ----------
    ser : serial.Serial() instance
        The device we are reading from.
    read_buffer : bytes, default b''
        Previous read buffer that is appended to.

    Returns
    -------
    output : bytes
        Bytes object that contains read_buffer + read.

    Notes
    -----
    .. `**args` appears, but is never used. This is for
       compatibility with `read_all_newlines()` as a
       drop-in replacement for this function.
    """
    # Set timeout to None to make sure we read all bytes
    previous_timeout = ser.timeout
    ser.timeout = None

    in_waiting = ser.in_waiting
    read = ser.read(size=in_waiting)

    # Reset to previous timeout
    ser.timeout = previous_timeout

    return read_buffer + read


def read_all_newlines(ser, read_buffer=b"", n_reads=4):
    """Read data in until encountering newlines.

    Parameters
    ----------
    ser : serial.Serial() instance
        The device we are reading from.
    n_reads : int
        The number of reads up to newlines
    read_buffer : bytes, default b''
        Previous read buffer that is appended to.

    Returns
    -------
    output : bytes
        Bytes object that contains read_buffer + read.

    Notes
    -----
    .. This is a drop-in replacement for read_all().
    """
    raw = read_buffer
    for _ in range(n_reads):
        raw += ser.read_until()
    return raw


def parse_read(read):
    """Parse a read with time, volage data

    Parameters
    ----------
    read : byte string
        Byte string with comma delimited time/voltage
        measurements.

    Returns
    -------
    voltage : list of floats; Voltages in volts.
    remaining_bytes : byte string remaining, unparsed bytes.
    """
    # Initiate Variables to Hold Voltages
    channelList = {0:[], 1:[], 2:[], 3:[]}
    time_ms = []

    # Separate independent time/voltage measurements
    pattern = re.compile(b"\d+|,|-")
    raw_list = [
        b"".join(pattern.findall(raw)).decode()
        for raw in read.split(b"\r\n")
    ]

    for raw in raw_list[:-1]:
        try:
            t, V1, V2, V3, v4 = raw.split(",")
            
            time_ms.append(int(t))
            for i,V in enumerate([V1, V2, V3, v4]):
                channelList[i].append(int(V) * 5/1023)
        except:
            pass
    
    """
    # Lopp Through Data and Extract the Voltagrs
    for rawRead in read:
        try:
            rawListSingleRead = re.findall(r'\d+', rawRead.decode())
            for raw in rawListSingleRead:
                # Takes in a number from 0 - 1023 and converts to Voltage out of 5
                rawFinger = 0
                for channelNum in range(numChannels):
                    if raw[rawFinger] == '1':
                        Voltage = int(raw[rawFinger:rawFinger + 4]) * 5/1023
                        rawFinger += 3
                    else:
                        Voltage = int(raw[rawFinger:rawFinger + 3]) * 5/1023
                        rawFinger += 3
                    channelList[channelNum].append(Voltage)
    
        except:
            pass
    """
    
    # Return the Voltages You Read in
    if len(read) == 0:
        return time_ms, channelList[0], channelList[1], channelList[2], channelList[3], b""
    else:
        return time_ms, channelList[0], channelList[1], channelList[2], channelList[3], b'' #raw_list[-1].encode()

def daq_stream_async(n_data, seeFullPlot, testNeuralNetwork = False, n_trash_reads=100, n_reads_per_chunk=400, delay=100, Controller=None, nn=None):
    """Obtain `n_data` data points from an Arduino stream
    with a delay of `delay` milliseconds between each."""
    print("Streaming in Data from the Arduino")
    global dataFinger
    
    # Find Arduino Port. If None: print Error
    try:
        port = find_arduino()
        arduino = serial.Serial(port, baudrate=115200)
    except Exception as e:
            print("No Port Found: ", e)
            exit    
    
    # Specify delay
    arduino.write(bytes([READ_DAQ_DELAY]) + (str(delay) + "x").encode())
    
    # Turn on the stream
    arduino.write(bytes([STREAM]))

    # Read and throw out first few reads
    for i in range(n_trash_reads):
        _ = arduino.read_until()

    # Receive data
    read_buffer = b""
    dataFinger = 0
    pointNum = 1
    t0 = time.time()
    while len(data["time_ms"]) < n_data:
        # Read in chunk of data
        raw = read_all_newlines(arduino, read_buffer=read_buffer, n_reads=n_reads_per_chunk)
        #print(len(data["time_ms"]))
        
        # Parse it, passing if it is gibberish
        try:
            time_ms, V1, V2, V3, V4, read_buffer = parse_read(raw)
            
            # Update data dictionary
            startNum = len(data["time_ms"])
            for i in range(len(time_ms)):
                data["time_ms"].append(startNum+i)
            data["Channel1"] += V1
            data["Channel2"] += V2
            data["Channel3"] += V3
            data["Channel4"] += V4
                        
            # When Ready, Send Data Off for Analysis
            pointNum = len(data["time_ms"])
            while pointNum - dataFinger >= xWidth:
                live_plotter(dataFinger, seeFullPlot, testNeuralNetwork=testNeuralNetwork, Controller=Controller, nn=nn)
                dataFinger += moveDataFinger
            
        except Exception as e:
            print(e)
            pass
        
    # Close the Arduino at the End
    tf = time.time()
    print(tf-t0)
    arduino.close()


# --------------------------------------------------------------------------- #
# --------------------- Extract Test Data from Excel ------------------------ #


def streamExcelData(testDataExcelFile, seeFullPlot = False, testNeuralNetwork = False, testSheetNum = 0, Controller=None, nn=None, analyzeSheet = None):
    """
    Extracts EMG Data from Excel Document (.xlsx). Data can be iplaced n any,
    worksheet which the user can specify using 'testSheetNum'.
    In the Worksheet:
        Channel 1 Data must be in Column 'A'
        Channel 2 Data must be in Column 'B'
        Channel 3 Data must be in Column 'C'
        Channel 4 Data must be in Column 'D'
    If No Data is present in a cell, it will be read in as zero.
    --------------------------------------------------------------------------
    Input Variable Definitions:
        testDataExcelFile: The Path to the Excel File Containing the Channel Data
        testSheetNum: An Integer Representing the Excel Worksheet (0-indexed) of the data.
    --------------------------------------------------------------------------
    """
    # Reset Data Variable in Case it Was Previously Populated
    global data
    data = dict(time_ms=[], Channel1=[], Channel2=[], Channel3=[], Channel4=[])  

    # Load Data from Excel File if No Sheet Given
    if analyzeSheet == None:
        print("\nAttempting to Extract Data from the Excel File:", testDataExcelFile)
        WB = xl.load_workbook(testDataExcelFile, data_only=True,read_only=True)
        WB_worksheets = WB.worksheets
        analyzeSheet = WB_worksheets[testSheetNum]
    
    # If Header Exists, Skip Until You Find the Data Data
    for row in analyzeSheet.rows:
        cellA = row[0]
        if type(cellA.value) == type(1.1):
            dataStartRow = cellA.row
            break
    
    # Open Movie Writer and Add Data
    #with writer.saving(fig, outputData + "EMG_Data.mp4", 300):
    dataFinger = 0
    # Loop Through the Excel Worksheet to collect all the data
    for pointNum, [colA, colB, colC, colD] in enumerate(analyzeSheet.iter_rows(min_col=1, min_row=dataStartRow, max_col=4, max_row=analyzeSheet.max_row)):
        # Get Cell Values for First 4 Channels: Represents the Voltage for Each Channel;
        V1 = colA.value; V2 = colB.value;
        V3 = colC.value; V4 = colD.value;
        
        # SafeGaurd: If User Edits the Document to Create Empty Rows, Stop Reading in Data
        if V1 == None and V2 == None and V3 == None and V4 == None:
            break
        
        # Add EMG Global Data to Dictionary in Sequential Order.
        data["time_ms"].append(pointNum)
        data["Channel1"].append(float(colA.value or 0))  # Represent No Value (None) as 0
        data["Channel2"].append(float(colB.value or 0))  # Represent No Value (None) as 0
        data["Channel3"].append(float(colC.value or 0))  # Represent No Value (None) as 0
        data["Channel4"].append(float(colD.value or 0))  # Represent No Value (None) as 0
        
        # When Ready, Send Data Off for Analysis
        while pointNum+1 - dataFinger >= xWidth:
            live_plotter(dataFinger, seeFullPlot, testNeuralNetwork, Controller=Controller, nn=nn)
            dataFinger += moveDataFinger
        
    # Finished Data Collection: Report Back to User
    print("\tDone Data Collecting from File: ", analyzeSheet.title)
    

def getDataFromTrainingFile(excelSheet, Training_Data, Training_Labels, movementOptions):
    # Get Hand Movement
    currentLabel = excelSheet.title.split(" - ")[1]
    featureLabel = movementOptions == currentLabel.lower()
    featureLabel = featureLabel.astype(int)
    
    # If Header Exists, Skip Until You Find the Data Data
    for row in excelSheet.rows:
        cellI = row[8]; cellJ = row[9]
        cellK = row[10]; cellL = row[11]
        if type(cellI.value) == type(1.1) or type(cellJ.value) == type(1.1) or type(cellK.value) == type(1.1) or type(cellL.value) == type(1.1):
            dataStartRow = row[0].row
            break
    
    # Create Data Structure to Hold Results Until Full Group is Present
    dataHold = np.empty((0, numChannels), float)
            
    # Loop Through the Excel Worksheet to collect all the data
    for (colI, colJ, colK, colL) in excelSheet.iter_rows(min_col=9, min_row=dataStartRow, max_col=9+numChannels-1, max_row=excelSheet.max_row):
        # Get Cell Values for First 4 Channels: Represents the Voltage for Each Channel;
        feature1 = colI.value; feature2 = colJ.value;
        feature3 = colK.value; feature4 = colL.value;
        
        # If All rows are empty, it is a new group
        if feature1 == None and feature2 == None and feature3 == None and feature4 == None:
            # Base Case: dataHold didnt change, we are done collecting
            if len(dataHold) == 0:
                break
            # Collect Point as Data
            featureAv = np.nanmean(np.where(dataHold>0, dataHold, np.nan), axis=0)
            Training_Data = np.vstack((Training_Data, featureAv))
            Training_Labels = np.vstack((Training_Labels, featureLabel))
            # Reset Group Holder
            dataHold = np.empty((0, numChannels), float)
            continue
        
        # Collect Features in Cirect Group
        featureList = [float(feature1 or 0), float(feature2 or 0), float(feature3 or 0), float(feature4 or 0)]
        dataHold = np.vstack((dataHold, featureList))
    
    print("\tCollected Training Data for:", excelSheet.title)
    return Training_Data, Training_Labels

def getTrainingData(trainDataExcelFolder, movementOptions, mode):
    if mode == 'Train':
        NN_Data = np.empty((0, numChannels), float)
        NN_Labels = np.empty((0, len(movementOptions)), float)
        
    for excelFile in list(os.listdir(trainDataExcelFolder)):
        if excelFile.endswith(".xlsx") and not excelFile.startswith("~"):
            # Get Full Path to the Excel File
            trainingExcelFile = trainDataExcelFolder + excelFile
            print("\nLoading Excel File", trainingExcelFile)
            # Load the Excel File
            if mode == 'Train':
                WB = xl.load_workbook(trainingExcelFile, data_only=True, read_only=True)
            else:
                WB = xl.load_workbook(trainingExcelFile, data_only=True, read_only=False)
            WB_worksheets = WB.worksheets
            # Loop Over Each Sheet in the File
            for excelSheet in WB_worksheets:
                resetGlobalVariables()
                
                # Get the Training Data/Label from the Sheet
                if mode == 'Train':
                    NN_Data, NN_Labels = getDataFromTrainingFile(excelSheet, NN_Data, NN_Labels, movementOptions)
                elif mode == 'reAnalyze':
                    print("\tReanalyzing Excel Sheet:", excelSheet.title)
                    # Delete the Previous Analysis from Excel
                    excelSheet.delete_cols(5,8) # Delete Columns E-L (5 -> 5+8-1)
                    # ReAnalyze Data (First Four Columns)
                    streamExcelData(trainingExcelFile, analyzeSheet=excelSheet)
                    # Overwrite Excel Sheet with new Analysis
                    handMovement = excelSheet.title.split(" - ")[1]
                    saveTestData(trainDataExcelFolder, excelFile, sheetName=excelSheet, handMovement=handMovement, overWrite = True)
                    
    if mode == 'Train':
        NaN_Placements = np.isnan(NN_Data)
        NN_Data[NaN_Placements] = 0
        return NN_Data, NN_Labels
    elif mode == 'reAnalyze':
        return None

def saveTestData(saveDataFolder, saveExcelName, sheetName = "Trial 1 - No Gesture", handMovement="No Movement", overWrite=False):
    # Get Data from Global Enviroment
    global data
    
    # Create Output File Directory to Save Data: If None
    os.makedirs(saveDataFolder, exist_ok=True)
    
    # Path to File to Save
    excel_file = saveDataFolder + saveExcelName
    
    # If the File is Not Present: Create The Excel File
    if not os.path.isfile(excel_file):
        print("\nSaving the Data as New Excel Workbook")
        # Make Excel WorkBook
        WB = xl.Workbook()
        WB_worksheet = WB.active 
        WB_worksheet.title = sheetName
    # Overwrite Previous Excel Sheet; Replacing Sheetname that was Edited
    elif overWrite:
        print("\tOverWriting Excel File:", excel_file)
        WB = xl.load_workbook(excel_file, read_only=False)
        WB_worksheet = sheetName
    # Loading in Previous Excel File, Creating New Sheet, Editing Trial Number in SheetName
    else:
        print("Excel File Already Exists. Loading File")
        WB = xl.load_workbook(excel_file, read_only=False)
        currentSheets = WB.sheetnames
        # Get All Sheets with the Current Movement
        currentMovementSheets = []
        for sheet in currentSheets:
            movement = sheet.split(" - ")
            if handMovement == movement[1]:
                currentMovementSheets.append(sheet)
        # Get the Last Trial for this Hand Movement
        sheetName = max(currentMovementSheets, key=lambda x: int(re.findall(r'\d+', x.split(" - ")[0])[0]), default= "Trial 0 - " + handMovement)
        # Edit SheetName
        sheetInfo = sheetName.split(" - ")
        currentTrial = re.findall(r'\d+', sheetInfo[0])[0]
        newTrial = sheetInfo[0].split(currentTrial)[0] + str(int(currentTrial)+1)
        sheetName = newTrial + " - " + sheetInfo[1]
        WB_worksheet = WB.create_sheet(sheetName)
        print("Saving Sheet as", sheetName)
    
    
    xPeakHeader = ['Channel 1 X Peaks', 'Channel 2 X Peaks', 'Channel 3 X Peaks', 'Channel 4 X Peaks']
    featureHeader = ['Channel 1 Features', 'Channel 2 Features', 'Channel 3 Features', 'Channel 4 Features']
    if not overWrite:
        # Label First Row
        header = ['Channel 1', 'Channel 2', 'Channel 3', 'Channel 4']
        header.extend(xPeakHeader)
        header.extend(featureHeader)
        WB_worksheet.append(header)
        
        # Save Data to Worksheet
        for dataNum in range(len(data['time_ms'])):
            row = []
            for channel in range(numChannels):
                row.append(data['Channel'+str(1+channel)][dataNum])
            WB_worksheet.append(row)
    else:
        for cells in WB_worksheet.iter_rows(min_col=5, min_row=1, max_col=12, max_row=1):
            colE, colF, colG, colH, colI, colJ, colK, colL = cells[0], cells[1], cells[2], cells[3], cells[4], cells[5], cells[6], cells[7]
            [colE.value, colF.value, colG.value, colH.value] = xPeakHeader
            [colI.value, colJ.value, colK.value, colL.value] = featureHeader

    
  #  for col in WB_worksheet.iter_cols(min_row=1, min_col = 1, max_col=3, max_row=1):
   #    print(col, len(col))
    #    col.sheetColsFeatures['A'].width = len(col)
        
    
    # Add X Peaks and Features
    sheetColsXPeaks = ['E','F','G','H']
    sheetColsFeatures = ['I','J','K','L']
    for channel in range(numChannels):
        startIndex = 1
        for peakNum in xTopGrouping[channel]:
            rowIndex = startIndex
            peakColor = (peakNum-1)%(len(openpyxlColors))
            cellColor = Color(indexed=openpyxlColors[peakColor])
            for peakVal in xTopGrouping[channel][peakNum]:
                WB_worksheet[sheetColsXPeaks[channel]][rowIndex].value = peakVal
                WB_worksheet[sheetColsXPeaks[channel]][rowIndex].fill = PatternFill(fgColor=cellColor, fill_type = 'solid')
                WB_worksheet[sheetColsXPeaks[channel]][rowIndex].alignment = Alignment(horizontal='center')
                rowIndex += 1
            # Set the Same Row Index for All
            startIndex += 1 + len(max(xTopGrouping.values(), key = lambda x: len(x[peakNum]))[peakNum])
            
    for channel in range(numChannels):
        startIndex = 1
        for peakNum in featureSetGrouping[channel]:
            rowIndex = startIndex
            peakColor = (peakNum-1)%(len(openpyxlColors))
            cellColor = Color(indexed=openpyxlColors[peakColor])
            for featureVal in featureSetGrouping[channel][peakNum]:
                WB_worksheet[sheetColsFeatures[channel]][rowIndex].value = featureVal
                WB_worksheet[sheetColsFeatures[channel]][rowIndex].fill = PatternFill(fgColor=cellColor, fill_type = 'solid')
                WB_worksheet[sheetColsFeatures[channel]][rowIndex].alignment = Alignment(horizontal='center')
                rowIndex += 1
            startIndex += 1 + len(max(xTopGrouping.values(), key = lambda x: len(x[peakNum]))[peakNum])
    
    # Save as New Excel File
    WB.save(excel_file)
    WB.close()
  
    # Reset Data Variable
    #data = dict(time_ms=[], Channel1=[], Channel2=[], Channel3=[], Channel4=[])  

def resetGlobalVariables():
    global data
    global lastPeakAnalyzed
    global currentGroupNum
    
    data = dict(time_ms=[], Channel1=[], Channel2=[], Channel3=[], Channel4=[])  
    
    lastPeakAnalyzed = 0
    currentGroupNum = -1
    
    for channelNum in range(numChannels):
        # Hold Analysis Values
        RMSDataList[channelNum] = []
        mapPeakLocToYPos[channelNum] = {}
        xTopGrouping[channelNum] = {1:[]}
        featureSetGrouping[channelNum] = {1:[]}


# --------------------------------------------------------------------------- #
# ------------------------------ Plot Data ---------------------------------- #

xDataEMG = data['time_ms'][0:xWidth]
def live_plotter(dataFinger, seeFullPlot = False, testNeuralNetwork = False, Controller=None, nn=None):
    global data
    global lastPeakAnalyzed
    global currentGroupNum
    global mapPeakLocToYPos
        
    # Get X Data: Shared Axis for All Channels
    xDataEMG = data['time_ms'][dataFinger:dataFinger+xWidth]
    # Add incoming Data to Each Respective Channel's Plot
    for dataChannel in range(numChannels):
        
        # ---------------------- Get EMG Signal -----------------------------#
        # Get New Y DataData
        newYData = data['Channel' + str(dataChannel+1)][dataFinger:dataFinger + xWidth]
        # Plot Raw EMG Data
        movieGraphChannelListRaw[dataChannel].set_data(xDataEMG, newYData)
        channelListRaw[dataChannel].set_xlim(xDataEMG[0], xDataEMG[0]+xWidth)
        # -------------------------------------------------------------------#
        
        # ---------------------- High pass Filter ---------------------------#
        # Calculate the Number of New Data Points You Need for Peak Detecting
        RMSData = RMSDataList[dataChannel]
        if len(RMSData) == 0 and dataFinger > 0:
            print("You Collected NO RMS Data Last Round ... ?")
            print("Please Decrease Your rmsWindow Size or Increase xWidth")
            sys.exit()
        else:
            newRMSDataBegins = stepSize*(len(RMSData)) - (dataFinger)

        # High Pass Filter to Remove Noise
        startHPF = max(dataFinger + newRMSDataBegins - badInitialHighPass,0)
        yDataBuffer = data['Channel' + str(dataChannel+1)][startHPF:dataFinger + xWidth]
        filteredData = highPassFilter(yDataBuffer, f1, f3, Rp, Rs, samplingFreq)[-xWidth+newRMSDataBegins:]   
        # -------------------------------------------------------------------#

        # --------------------- Root Mean Squared ---------------------------#
        if len(filteredData) >= rmsWindow:
            newRMSData = RMSFilter(filteredData, rmsWindow, stepSize)
            RMSData.extend(newRMSData)
        else:
            print("If You are Here, it is a Mistake")
            print("Please Decrease Your rmsWindow Size or Increase xWidth")
            sys.exit()
        xDataRMS = data['time_ms'][max(len(RMSData) - xWidthPeaks,0):len(RMSData)]  
        # Plot RMS Data
        movieGraphChannelListFiltered[dataChannel].set_data(xDataRMS, RMSData[-xWidthPeaks:])
        channelListFiltered[dataChannel].set_xlim(xDataRMS[0], xDataRMS[0] + xWidthPeaks)
        # -------------------------------------------------------------------#
        
        # ------------------------- Move Robot  -----------------------------#
        # If Testing the Robot, Send Peaks Through NN After Peak is Collected
        if testNeuralNetwork:
            currentPeakAnalyzing = xDataRMS[-(len(newRMSData) + peakDetectionBufferSize)]
            lastXPeakLoc = max([max(max(xTopGrouping[dataChannel].values()), default=0) for i in range(numChannels)])
            if currentPeakAnalyzing - lastXPeakLoc > maxPeakSep and lastPeakAnalyzed < currentGroupNum:
                lastPeakAnalyzed = currentGroupNum
                groupFeatures = []
                for channel in range(numChannels):
                    channelFeatures = xTopGrouping[channel][currentGroupNum]
                    if channelFeatures == []:
                        groupFeatures.append(0)
                    else:
                        groupFeatures.append(np.mean(channelFeatures))
                inputData = np.array([groupFeatures])
                predictionProbs = nn.model.predict(inputData)
                predictedIndex = np.argmax(predictionProbs)
                predictedLabel = movementOptions[predictedIndex]
                print("The Predicted Label is", predictedLabel)
                if Controller:
                    if predictedLabel == "left":
                        Controller.moveLeft()
                    elif predictedLabel == "right":
                        Controller.moveRight()
                    elif predictedLabel == "down":
                        Controller.moveDown()
                    elif predictedLabel == "up":
                        Controller.moveUp()
                    elif predictedLabel == "grab":
                        print("Grabbing")
                        continue
                    elif predictedLabel == "release":
                        print("Releasing")
                        continue
        # -------------------------------------------------------------------#
        
        # ----------------------- Peak Detection  ---------------------------#
        # Get Most Current RMS Data (Add Buffer in Case the peak is Cut Off)
        bufferRMSData = RMSData[-(len(newRMSData) + peakDetectionBufferSize):]
        bufferRMSDataX = xDataRMS[-(len(newRMSData) + peakDetectionBufferSize):]
        # Find Peaks from the New Data
        newTopPeaks, mapPeakLocToYPos = find_peaks(bufferRMSDataX, bufferRMSData, mapPeakLocToYPos, dataChannel, RMSData)
        # If No New Peaks, Then No New Features
        if newTopPeaks == {}:
            continue
        # Split the Peaks into the X,Y Points
        xPeakTop, yPeakTop = zip(*newTopPeaks.items())
        # -------------------------------------------------------------------#
        
        # --------------------- Feature Extraction  -------------------------#
        # Features Analysis to Group Peaks Together  
        currentGroupNum = max(xTopGrouping[dataChannel].keys(), default=0)
        #currentGroup = xTopGrouping[dataChannel].get(currentGroupNum, [])
        highestXPeak = max([max(max(xTopGrouping[dataChannel].values()), default=0) for i in range(numChannels)])
        if abs(xPeakTop[0] - highestXPeak) > maxPeakSep and highestXPeak != 0:
            currentGroupNum += 1
            for channel in range(numChannels):
                xTopGrouping[channel][currentGroupNum] = []
                featureSetGrouping[channel][currentGroupNum] = []
        batchXGroups, featureSetTemp = featureDefinition(RMSData, xPeakTop, currentGroupNum)
        # Get New Groups
        #newXGroupsNumbers = xTopGrouping[dataChannel].keys() - batchXGroups.keys()
        # Update Overall Grouping Dictionary
        for groupNum in batchXGroups.keys():
            # Get New Peaks/Features to Add
            updateXGroups = batchXGroups[groupNum]
            updateFeatures = featureSetTemp[groupNum]
            # Add Them
            if groupNum in xTopGrouping[dataChannel].keys():
                xTopGrouping[dataChannel][groupNum].extend(updateXGroups)
                featureSetGrouping[dataChannel][groupNum].extend(updateFeatures)
            else:
                xTopGrouping[dataChannel][groupNum] = updateXGroups
                featureSetGrouping[dataChannel][groupNum] = updateFeatures
                        
        
        # Plot the Peaks; Colored by Grouping
        yFinger = 0
        yPeakTop = list(mapPeakLocToYPos[dataChannel].values())
        for groupNum in xTopGrouping[dataChannel].keys():
            # Check to See if the Group Has a Plot You Can Use
            groupPeakPlot = movieGraphChannelTopPeaksList[dataChannel].get(groupNum, None)
            # If None Availible, Create a New Plot to Add the Data
            if groupPeakPlot == None:
                channelFiltered = channelListFiltered[dataChannel]
                # Color Code the Group Peaks. Wrap Around to First Index When Done
                groupColor = (groupNum-1)%(len(peakCurrentRightColorOrder))
                # Create a Plot for the Peaks Using its Respective Group's Color
                groupPeakPlot = channelFiltered.plot([], [], 'o', c=peakCurrentRightColorOrder[groupColor], linewidth=1, alpha = 0.65)[0]
                # Save the Plot for Later Use in the Group
                movieGraphChannelTopPeaksList[dataChannel][groupNum] = groupPeakPlot
            # Get Peak Points
            xTopGroup = xTopGrouping[dataChannel][groupNum]
            yTopGroup = yPeakTop[yFinger:yFinger+len(xTopGroup)]
            # Add the Data to the Plot
            xTopGroupNumpy = np.array(xTopGroup)
            xTopGroupPlotWindow = xTopGroupNumpy[xTopGroupNumpy > xDataRMS[0]]
            yTopGroupPlotWindow = yTopGroup[-len(xTopGroupPlotWindow):]
            if len(xTopGroupPlotWindow) == 0:
                yTopGroupPlotWindow = [] 
            groupPeakPlot.set_data(xTopGroupPlotWindow, yTopGroupPlotWindow)
            # Keep Track of Current Y Data with Respect to the X Group Data
            yFinger += len(xTopGroup)
        
        #print("\n")
        
    #writer.grab_frame() 
    try:
        # Update to Get New Data Next Round
        #plt.show(block=False)
        plt.draw()
        fig.canvas.flush_events()
                 
        # Write to Video
        
    except Exception as e:
        print(e)
        sys.exit()
    

def matchMatlabPlots():
    print("Printing Seperate test plots")
    # Wait to get all the points First
    totalPoints1 = 0; totalPoints2 = 1;
    while totalPoints1 != totalPoints2:
        totalPoints1 = len(data['Channel1'])
        time.sleep(1)
        totalPoints2 = len(data['Channel1'])
        
    xData = data['time_ms']
    yData = data['Channel1']
    
    f1 = 100; f3 = 50;
    Rp = 0.1; Rs = 30;
    samplingFreq = 1000
    window = 250; step = 8;
    print("Ready to Plot Data")
    
    # Get Data and Filter
    plt.figure()
    plt.plot(xData,yData, c='tab:blue', alpha=0.7)
    plt.title("EMG Data")
    
    plt.figure()
    filteredData = highPassFilter(yData, f1, f3, Rp, Rs, samplingFreq)
    plt.plot(xData,filteredData, c='tab:blue', alpha=0.7)
    plt.title("Filtered Data")
    
    plt.figure()
    RMSData = RMSFilter(filteredData, window, step)
    plt.plot(xData[0:len(RMSData)],RMSData, c='tab:blue', alpha=0.7)
    plt.title("RMS Data")
    
    # Find Peaks
    batchTopPeaks = find_peaks(xData, RMSData)
    xPeakTop, yPeakTop = zip(*batchTopPeaks.items())
    xTopGrouping, featureSet = featureDefinition(RMSData, xPeakTop, 0)
    
    """
    yFinger = 0
    for groupNum in xTopGrouping.keys():
        xTopGroup = xTopGrouping[groupNum]
        yTopGroup = yTop[yFinger:yFinger+len(xTopGroup)]
        plt.plot(xTopGroup, yTopGroup, "*", c=peakCurrentRightColorOrder[groupNum])
        yFinger += len(xTopGroup)
    """

        


# --------------------------------------------------------------------------- #
# ------------------------- Signal Analysis --------------------------------- #


def highPassFilter(inputData, f1, f3, Rp, Rs, samplingFreq):
    """
    data: Data to Filter
    f1: cutOffFreqPassThrough
    f3: cutOffFreqBand
    Rp: attDB (0.1)
    Rs: cutOffDB (30)
    samplingFreq: Frequecy You Take Data
    """
    Wp = 2*math.pi*f1/samplingFreq
    Ws = 2*math.pi*f3/samplingFreq
    [n, wn] = scipy.signal.cheb1ord(Wp/math.pi, Ws/math.pi, Rp, Rs)
    [bz1, az1] = scipy.signal.cheby1(n, Rp, Wp/math.pi, 'High')
    filteredData = lfilter(bz1, az1, inputData)
    return filteredData

def RMSFilter(inputData, rmsWindow=250, stepSize=8):
    """
    The Function loops through the given EMG Data, looking at batches of data
        of size rmsWindow at every interval seperated by stepSize.
    In Each Window, we take the magnitude of the data vector (sqrt[a^2+b^2]
        for [a,b] data point)
    A list of each root mean squared value is returned (in order)
    
    The Final List has a length of 1 + math.floor((len(inputData) - rmsWindow) / stepSize)
    --------------------------------------------------------------------------
    Input Variable Definitions:
        inputData: A List containing the  EMG Data
        rmsWindow: The Amount of Data in the Groups we Analyze via RMS
        stepSize: The Distance Between Data Groups
    --------------------------------------------------------------------------
    """
    normalization = math.sqrt(rmsWindow)
    # Take Root Mean Squared of Batch Data (numBatch = rmsWindow)
    numSteps = 1 + math.floor((len(inputData) - rmsWindow) / stepSize)
    #print("RMS:", numSteps, len(inputData))
    RMSData = np.zeros(numSteps)
    for i in range(numSteps):
        # Get Data in the Window to take RMS
        inputWindow = inputData[i*stepSize:i*stepSize + rmsWindow]
        # Take RMS
        RMSData[i] = np.linalg.norm(inputWindow, ord=2)/normalization
    
    return RMSData

def featureDefinition(RMSData, xTop, currentGroupNum):
    featureSet = []
    correctionTerm = 0
    # For Every Peak, Take the Average of the Points in Front of it
    for xTopLoc in xTop:
        # Get Peak's Points
        featureWindow = RMSData[xTopLoc-featureDefinitionBuffer:xTopLoc+featureDefinitionBuffer]
        # Take the Average of the Peak as the Feature
        if len(featureWindow) > 0:
            featureSet.append(np.mean(featureWindow))
        # Edge Effect: np.mean([]) = NaN -> Ignore This Peak Until Next Round
        else:
            correctionTerm += 1
    
    # If No Features/Peaks This Round, Return Empty Dictionaries
    if featureSet == []:
        return {}, {}
        
    # Group the Feature Sets Peaks into One EMG Signal/Group
    peakSeperation = np.diff(xTop[0:len(xTop) - correctionTerm]) # Identify New Signal by Peak Seperation
    peakGrouping = {currentGroupNum:[featureSet[0]]}  # Holder for the Features
    xGrouping = {currentGroupNum:[xTop[0]]}        # Holder for the Corresponding Peaks
    for i, peakSep in enumerate(peakSeperation):
        # A Part of Previous Group
        if peakSep < maxPeakSep:
            peakGrouping[currentGroupNum].append(featureSet[i+1])
            xGrouping[currentGroupNum].append(xTop[i+1])
        # New Group
        else:
            currentGroupNum += 1
            for channel in range(numChannels):
                xTopGrouping[channel][currentGroupNum] = []
                featureSetGrouping[channel][currentGroupNum] = []
            peakGrouping[currentGroupNum] = [featureSet[i+1]]
            xGrouping[currentGroupNum] = [xTop[i+1]]
 
    return xGrouping, peakGrouping


def find_peaks(xData, yData, oldPeaks, channel, RMSData):
    # Convert to Numpy (For Faster Data Processing)
    numpyDataY = np.array(yData)
    numpyDataX = np.array(xData)
    # Find Peak Indices
    indicesTop = scipy.signal.find_peaks(yData, prominence=.05, height=0.01, distance=100)[0]
    # Get X,Y Peaks
    xTop = numpyDataX[indicesTop]
    yTop = numpyDataY[indicesTop]
    # Find the New Peaks
    newTopPeaks = {}
    for i, xLoc in enumerate(xTop):
        if xLoc not in oldPeaks[channel] and xLoc + featureDefinitionBuffer > len(RMSData):
            # Record New Peaks and Add New Peaks to Ongoing list
            newTopPeaks[xLoc] = yTop[i]
            oldPeaks[xLoc] = yTop[i]
    # Return New Peaks and Update Peak Dictionary
    return newTopPeaks, oldPeaks


if __name__ == "__main__":
    # ---------------------------------------------------------------------- #
    #    User Parameters to Edit (More Complex Edits are Inside the Files)   #
    # ---------------------------------------------------------------------- #
    
    # General Data Collection Information
    numDataPoints = 60000  # Number of Points to Stream into the Arduino
    seeFullPlot = True    # (NOT IMPLEMENTED YET) Graph the Peak Analysis IN ADDITION TO the Arduino Data
    SaveNeuralNetwork = False
    testNeuralNetwork = False
    optimizeNNParams = False
    
    # User Options (Only One Can be True; Only the First True Variable Excecutes)
    streamArduino = False          # Stream in Data from the Arduino
    readSignalsFromExcel = False  # Uses the test Data Provided in testDataExcelFile on Sheet testSheetNum
    trainNeuralNetwork = True
    reAnalyzeTrainingPeaks = False
    
    # Take Data from the Arduino and Save it as an Excel (For Later Use)
    if streamArduino or readSignalsFromExcel:
        saveInputData = False                              # Saves the Data Streamed into From the Arduino as saveExcelName
        saveExcelName = "Samuel Solomon 2021-02-24.xlsx"  # The Name of the Saved File
        saveDataFolder = "../Input Data/Training Data/"   # Data Folder to Save the Excel Data; MUST END IN '/'
        handMovement = "Right"                 # Speficy the hand Movement You Will Perform
    else:
        saveInputData = False
    
    # Instead of Arduino Data, Use Test Data from Excel File
    if readSignalsFromExcel:
        testDataExcelFile = "../Input Data/Test Data/channel.xlsx" # Path to the Test Data
        testSheetNum = 0                     # The Sheet/Tab Order (Zeroth/First/Second/Third) on the Bottom of the Excel Document
    
    # Use Previously Processed Data that was Saved; Extract Features for Training
    if reAnalyzeTrainingPeaks or trainNeuralNetwork:
        trainDataExcelFolder = "../Input Data/Training Data/"  # Path to the Training Data Folder; All .xlsx Data Used
    
    if trainNeuralNetwork or testNeuralNetwork:
        neuralNetworkName = "V1"
        neuralNetworkFolder = "../Neural Network/"
    
    # Saving the Neural network
    if SaveNeuralNetwork:
        saveNeuralNetworkName = "V1"
        saveNeuralNetworkFolder = "./Neural Network/"
    
    # ---------------------------------------------------------------------- #
    #                    Parameters User SHOULD NOT Edit                     #
    # ---------------------------------------------------------------------- #
    
    # Variables Users Can Change, BUT SHOULDNT
    if streamArduino or readSignalsFromExcel:
        sheetName = "Trial 1 - "  # If SheetName Already Exists, Excel 1 to Trial #
        sheetName = sheetName + handMovement
    
    # ---------------------------------------------------------------------- #
    # ---------------------------------------------------------------------- #
    # ---------------------------------------------------------------------- #
    #           Initiate Neural Network (Should Not Have to Edit)            #
    # ---------------------------------------------------------------------- #
    # ---------------------------------------------------------------------- #
    # ---------------------------------------------------------------------- #
    
    # Define Labels as Array
    movementOptions = np.array(["Right", "Left", "Up", "Down", "Grab", "Release"])
    movementOptions = np.char.lower(movementOptions)
    # Edge Case: User Defines a Movement that Does not Exist, Return Error
    if saveInputData and handMovement.lower() not in movementOptions:
        print("\nUser Defined an Unknown Hand Gesture")
        print("The Gesture", "'" + handMovement.lower() + "'", "is Not in", movementOptions)
        sys.exit()
    
    if trainNeuralNetwork or testNeuralNetwork:
        outputNeuralNetwork = neuralNetworkFolder + neuralNetworkName
        if optimizeNNParams:
            optimizeNN = NeuralNet.Helpers(name = outputNeuralNetwork, dataDimension = numChannels, numClasses = len(movementOptions))
            neuralNetList = optimizeNN.neuralPermutations()
        # Make the Neural Network  (dim = The dimensionality of one data point) 
        else:
            neuralNetList = [NeuralNet.Neural_Network(name = outputNeuralNetwork, dim = numChannels)]
            nn = neuralNetList[0]
    else:
        nn = None
        
    # ---------------------------------------------------------------------- #
    # ---------------------------------------------------------------------- #
    # ---------------------------------------------------------------------- #
    #           Data Collection Program (Should Not Have to Edit)            #
    # ---------------------------------------------------------------------- #
    # ---------------------------------------------------------------------- #
    # ---------------------------------------------------------------------- #
          
    # Stream in Data from Arduino (EMG Signals)
    if streamArduino:
        daq_stream_async(numDataPoints, seeFullPlot, testNeuralNetwork, nn=nn)
    # Take Data from Excel Sheet
    elif readSignalsFromExcel:
        streamExcelData(testDataExcelFile, seeFullPlot, testNeuralNetwork, testSheetNum, nn=nn)
    # Take Preprocessed (Saved) Features from Excel Sheet
    elif trainNeuralNetwork:
        signalData, signalLabels = getTrainingData(trainDataExcelFolder, movementOptions, mode='Train')
        print("\nCollected Signal Data")
    elif reAnalyzeTrainingPeaks:
        getTrainingData(trainDataExcelFolder, movementOptions, mode='reAnalyze')
    
    # Train the ML
    if trainNeuralNetwork:
        # Split the Data into Training and Validation Sets
        Training_Data, Testing_Data, Training_Labels, Testing_Labels = train_test_split(signalData, signalLabels, test_size=0.33, shuffle= True, stratify=signalLabels)
        # Find Best Neural Network
        goodNets = {}
        for nn in neuralNetList:
            try:  
                # Train the NN with the Training Data
                Neural_Network_Statistics = nn.train_model(Training_Data, Training_Labels, 300, seeTrainingSteps = True)
                # Make a prediction using new data
                nn.neural_net_prediction(Testing_Data, Testing_Labels)
                # Plot the training loss    
                nn.plot_statistics(Neural_Network_Statistics)
                
                # Print out the Predictions from the Validation Data to Check
                numRight = 0
                posWrong = [0,0,0,0,0,0]
                predictions = nn.model.predict(Testing_Data)
                for i,testPoint in enumerate(predictions):
                    
                    testingIndex = np.argmax(testPoint)
                    testingLabel = movementOptions[testingIndex]
                    
                    realIndex = np.argmax(Testing_Labels[i])
                    realLabel = movementOptions[realIndex]
                    
                    #print("Input:", realLabel, "; NN Label:", testingLabel)
                    numRight += realLabel == testingLabel
                    if realLabel != testingLabel:
                        posWrong[realIndex] += 1
                accuracy = numRight/(i+1)
                print("\nAccuracy: ", accuracy)
                print(posWrong)
                print(movementOptions)
                goodNets[(accuracy, nn.opt, nn.loss, nn.metric)] = accuracy
            except Exception as e:
                print("\n", e)
                print((nn.opt, nn.loss, nn.metric))
        
    # Save the Data in Excel: EMG Channels (Cols 1-4); X-Peaks (Cols 5-8); Peak Features (Cols 9-12)
    if saveInputData:
        print("\nPeaks:", xTopGrouping)
        verifiedSave = input("Are you Sure you Want to Save the Data (Y/N): ")
        if verifiedSave.upper() == "Y":
            saveTestData(saveDataFolder, saveExcelName, sheetName, handMovement)
        else:
            print("User Chose Not to Save the Data")
    
    # Save the Neural Network (The Weights of Each Edge)
    if SaveNeuralNetwork and trainNeuralNetwork:
        outputNueralNetwork = saveNeuralNetworkName + saveNeuralNetworkName
        nn.save_model(outputNueralNetwork)


 